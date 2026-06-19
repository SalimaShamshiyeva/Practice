"""Step 6 - Generate the quotation spreadsheet.

If a company template (.xlsx) is supplied, rows are appended under its header
row (matched by column name, in any language). Otherwise a clean, styled
workbook is produced with the required columns. Status cells are color-coded
green / yellow / red.
"""

from __future__ import annotations

from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import MatchResult, MatchStatus
from .text_utils import normalize_header


# Output columns in order. The header text on the left is what we write / look
# for in a template; `field` is how we pull the value from a MatchResult.
COLUMNS = [
    ("Client Product", "client_product"),
    ("Quantity", "quantity"),
    ("Unit", "unit"),
    ("Analysis Result", "analysis_result"),
    ("Status", "status"),
    ("Warehouse Quantity", "warehouse_quantity"),
    ("Cost Price", "cost_price"),
    ("Supplier", "supplier"),
    ("Comments", "comments"),
]

_FILL = {
    MatchStatus.ON_STOCK: PatternFill("solid", fgColor="C6EFCE"),
    MatchStatus.PREVIOUSLY_PURCHASED: PatternFill("solid", fgColor="FFEB9C"),
    MatchStatus.NEW_PRODUCT: PatternFill("solid", fgColor="FFC7CE"),
}
_FONT_COLOR = {
    MatchStatus.ON_STOCK: "006100",
    MatchStatus.PREVIOUSLY_PURCHASED: "9C6500",
    MatchStatus.NEW_PRODUCT: "9C0006",
}


def _row_values(m: MatchResult) -> dict:
    p = m.product
    return {
        "client_product": p.name + (f" — {p.description}" if p.description else ""),
        "quantity": p.quantity if p.quantity is not None else "",
        "unit": p.unit,
        "analysis_result": m.matched_name or "—",
        "status": f"{m.status.emoji} {m.status.value}",
        "warehouse_quantity": (
            m.warehouse_quantity if m.warehouse_quantity is not None else ""
        ),
        "cost_price": m.cost_price if m.cost_price is not None else "",
        "supplier": m.supplier,
        "comments": m.comments,
    }


class ExcelGenerator:
    def generate(
        self,
        matches: list[MatchResult],
        output_path: str,
        template_path: Optional[str] = None,
    ) -> str:
        if template_path:
            return self._fill_template(matches, output_path, template_path)
        return self._new_workbook(matches, output_path)

    # ------------------------------------------------------------------ #
    def _new_workbook(self, matches, output_path) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Quotation"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="305496")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for col_idx, (title, _) in enumerate(COLUMNS, start=1):
            c = ws.cell(row=1, column=col_idx, value=title)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border

        for r, m in enumerate(matches, start=2):
            vals = _row_values(m)
            for col_idx, (_, field) in enumerate(COLUMNS, start=1):
                c = ws.cell(row=r, column=col_idx, value=vals[field])
                c.border = border
                c.alignment = center if field in (
                    "quantity", "unit", "status", "warehouse_quantity", "cost_price"
                ) else left
                if field == "status":
                    c.fill = _FILL[m.status]
                    c.font = Font(bold=True, color=_FONT_COLOR[m.status])

        self._autosize(ws)
        ws.freeze_panes = "A2"
        wb.save(output_path)
        return output_path

    # ------------------------------------------------------------------ #
    def _fill_template(self, matches, output_path, template_path) -> str:
        wb = load_workbook(template_path)
        ws = wb.active

        header_row, col_index = self._locate_header(ws)
        if header_row is None:
            # template has no recognizable header -> just use a fresh sheet layout
            return self._new_workbook(matches, output_path)

        start = header_row + 1
        # find first empty data row at/after start
        while any(
            ws.cell(row=start, column=c).value not in (None, "")
            for c in col_index.values()
        ):
            start += 1

        for offset, m in enumerate(matches):
            r = start + offset
            vals = _row_values(m)
            for field, col in col_index.items():
                ws.cell(row=r, column=col, value=vals.get(field, ""))
            status_col = col_index.get("status")
            if status_col:
                cell = ws.cell(row=r, column=status_col)
                cell.fill = _FILL[m.status]

        wb.save(output_path)
        return output_path

    def _locate_header(self, ws):
        """Find the header row and map our fields onto its columns by name."""
        # Map possible header texts -> our field keys.
        alias = {
            "client product": "client_product", "product": "client_product",
            "наименование": "client_product", "товар": "client_product",
            "quantity": "quantity", "qty": "quantity", "количество": "quantity",
            "кол-во": "quantity",
            "unit": "unit", "ед": "unit", "единица": "unit",
            "analysis result": "analysis_result", "результат": "analysis_result",
            "status": "status", "статус": "status",
            "warehouse quantity": "warehouse_quantity", "остаток": "warehouse_quantity",
            "склад": "warehouse_quantity",
            "cost price": "cost_price", "цена": "cost_price", "стоимость": "cost_price",
            "supplier": "supplier", "поставщик": "supplier",
            "comments": "comments", "комментарий": "comments", "примечание": "comments",
        }
        max_scan = min(ws.max_row, 15)
        for row in range(1, max_scan + 1):
            mapping = {}
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=row, column=col).value
                if val is None:
                    continue
                key = normalize_header(str(val))
                if key in alias:
                    mapping[alias[key]] = col
            if len(mapping) >= 3:  # looks like a header row
                return row, mapping
        return None, {}

    @staticmethod
    def _autosize(ws):
        widths = {1: 42, 2: 10, 3: 8, 4: 32, 5: 22, 6: 18, 7: 12, 8: 22, 9: 40}
        for idx, w in widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = w
